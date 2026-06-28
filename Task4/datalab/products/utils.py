import io,os,uuid
import pandas as pd
import numpy as np

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,PatternFill
from openpyxl.utils import get_column_letter

def read_any(file_path,sheet_name=None):
    if file_path.lower().endswith(('.xlsx','.xls')):
        return pd.read_excel(file_path,sheet_name=sheet_name or 0)
    return pd.read_csv(file_path)

def clean_columns(df:pd.DataFrame) -> pd.DataFrame:
    df.columns = (df.columns
                  .str.strip().str.replace(' ','_')
                  .str.replace(r"[^0-9a-zA-Z_]","",regex=True).str.lower())
    return df

def coerce(df,col,numeric=True):
    if col in df.columns:
        if numeric:
            df[col] = pd.to_numeric(df[col],errors='coerce')
        else:
            df[col]=df[col].astype(str).str.strip()
    return df

def normalize_for_product(df:pd.DataFrame) -> pd.DataFrame:
    '''
        Required Columns
        sku,name,category,price,quantity,tx_date
    '''

    df=clean_columns(df)

    rename_map={
        "product_sku":"sku","product":"name","title":"name",
        "cat":"category","qty":"quantity","date":"tx_date",
    }

    df=df.rename(columns=rename_map)

    for c in ["price","quantity"]:
        df=coerce(df,c,True)
    df=coerce(df,"sku",numeric=False)
    df=coerce(df,"name",numeric=False)
    df=coerce(df,"category",numeric=False)

    if "tx_date" in df.columns:
        df["tx_date"]=pd.to_datetime(df["tx_date"],errors='coerce').dt.date

    df=df.dropna(subset=["sku","name","price","quantity","tx_date"])

    df["quantity"]=df["quantity"].clip(lower=0)
    df["price"]=df["price"].clip(lower=0)

    return df

def df_to_excel_response(df:pd.DataFrame,fname="export.xlsx"):
    out_dir=os.path.join(settings.MEDIA_ROOT,"exports")
    os.makedirs(out_dir,exist_ok=True)
    fpath = os.path.join(out_dir, f"{uuid.uuid4().hex}_{fname}")
    with pd.ExcelWriter(fpath,engine="openpyxl") as w:
        df.to_excel(w,index=False,sheet_name="Products")

    wb=load_workbook(fpath)
    ws=wb.active

    header_font=Font(size=14,bold=True,color="FFFFFF")
    header_fill=PatternFill(start_color="4F81BD",end_color="4F81BD",fill_type="solid")


    blue_fill=PatternFill(start_color="BDD7EE",end_color="BDD7EE",fill_type="solid")
    green_fill=PatternFill(start_color="C6E0B4",end_color="C6E0B4",fill_type="solid")

    for col_num,col_name in enumerate(df.columns,1):
        cell=ws.cell(row=1,column=col_num)
        cell.font=header_font
        cell.alignment=Alignment(horizontal='center',vertical='center')

        cell.fill=header_fill
        ws.column_dimensions[get_column_letter(col_num)].width = 30

        if col_name.lower()=="sku":
            cell.fill=blue_fill
        elif col_name.lower()=="price":
            cell.fill=green_fill

    for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=ws.max_column):
        for cell in row:
            col_header=ws.cell(row=1,column=cell.column).value
            if str(col_header).lower()=="sku":
                cell.fill=blue_fill
            elif str(col_header).lower()=="price":
                cell.fill=green_fill

    wb.save(fpath)
    return fpath
